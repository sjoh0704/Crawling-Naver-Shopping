from openpyxl import Workbook
import os
import time




def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
            # print("파일생성")
    except OSError:
        print('Error: Creating directory. ' + directory)




def make_file(item_list):
        # 엑셀파일 쓰기
    write_wb = Workbook()

        # 이름이 있는 시트를 생성
    write_ws = write_wb.create_sheet('Naver Shopping')

        # Sheet1에다 입력
    sheet = write_wb.active
    # write_ws['A1'] = '숫자'

        #행 단위로 추가
    # write_ws.append([1,2,3])



        #셀 단위로 추가
    # write_ws.cell(5, 5, '5행5열')
    sheet.column_dimensions['A'].width = 100
    for i in range(len(item_list)):


        sheet.cell(i+1, 1).value = item_list[i].title
        sheet.cell(i+1, 1).hyperlink = item_list[i].url
        # sheet.cell(i + 1, 1).value.hyperlink = item_list[i].url

    createFolder('/NaverShopping')
    # write_wb.save(r"C:\Users\SeungJu\Desktop\pro_excel/navershopping.xlsx")
    now = time.strftime('%y%m%d-%H%M%S')

    write_wb.save(r"C:\NaverShopping\{}.xlsx".format(now))
